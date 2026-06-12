from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Agent ID", "Agent Name", "Creator Type",
    "Active Users (Licensed)", "Active Users (Unlicensed)",
    "Responses Sent", "Last Activity Date",
]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— No M365 Usage agent data imported (set M365USAGE_REPORT_Agents) —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r.get("agent_id", ""))
            ws.cell(row=i, column=2, value=r.get("agent_name", ""))
            ws.cell(row=i, column=3, value=r.get("creator_type", ""))
            ws.cell(row=i, column=4, value=r.get("active_users_licensed"))
            ws.cell(row=i, column=5, value=r.get("active_users_unlicensed"))
            ws.cell(row=i, column=6, value=r.get("responses_sent"))
            ws.cell(row=i, column=7, value=r.get("last_activity_date", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
