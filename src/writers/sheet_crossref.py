from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.crossref import FLAG_BOTH, FLAG_CLEAN, FLAG_MONITOR_ONLY, FLAG_OTEL_ONLY
from src.writers._style import autofit_columns, write_headers

HEADERS = [
    "Conversation ID",
    "Agent ID",
    "Agent Name",
    "Flag",
    "OTel DLP Signal",
    "AZ Exception Type",
    "AZ Alert Name",
    "Invocation Time",
    "Env ID",
]

_FILLS = {
    FLAG_BOTH:         PatternFill("solid", fgColor="FFCCCC"),   # red
    FLAG_MONITOR_ONLY: PatternFill("solid", fgColor="FFE5CC"),   # orange
    FLAG_OTEL_ONLY:    PatternFill("solid", fgColor="FFFACC"),   # yellow
    FLAG_CLEAN:        PatternFill("solid", fgColor="CCFFCC"),   # green
}


def write(ws: Worksheet, crossref_summary: list[dict]) -> None:
    write_headers(ws, HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    if not crossref_summary:
        ws.cell(row=2, column=1, value="— No flagged conversations (all clean or no Azure Monitor data) —")
        autofit_columns(ws, HEADERS)
        return

    for row_idx, row in enumerate(crossref_summary, start=2):
        flag = row.get("flag", FLAG_CLEAN)

        ws.cell(row=row_idx, column=1, value=row.get("conversation_id", ""))
        ws.cell(row=row_idx, column=2, value=row.get("agent_id", ""))
        ws.cell(row=row_idx, column=3, value=row.get("agent_name", ""))
        ws.cell(row=row_idx, column=4, value=flag)
        ws.cell(row=row_idx, column=5, value="Yes" if row.get("otel_dlp_signal") else "No")
        ws.cell(row=row_idx, column=6, value=row.get("az_exception_type", ""))
        ws.cell(row=row_idx, column=7, value=row.get("az_alert_name", ""))
        ws.cell(row=row_idx, column=8, value=row.get("invocation_time", ""))
        ws.cell(row=row_idx, column=9, value=row.get("env_id", ""))

        fill = _FILLS.get(flag, _FILLS[FLAG_CLEAN])
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).fill = fill

    autofit_columns(ws, HEADERS)
