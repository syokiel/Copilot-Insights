from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import write_headers

HEADERS = [
    "Person ID", "Metric Date", "Organization", "Is Active",
    "Total Actions", "Active Days", "Enabled Days",
    "Meeting Hours", "Uninterrupted Hours", "Multitasking Hours",
    "Conflicting Meeting Hours", "Small Meeting Hours",
    "Chats Sent", "Emails Sent", "Emails Sent w/ Copilot",
    "Meeting Hours Summarized", "Meetings Summarized",
    "Chat Web Prompts", "Chat Work Prompts",
]

_FIELDS = [
    "person_id", "metric_date", "organization", "is_active",
    "total_copilot_actions", "total_copilot_active_days", "total_copilot_enabled_days",
    "meeting_hours", "uninterrupted_hours", "multitasking_hours",
    "conflicting_meeting_hours", "small_meeting_hours",
    "chats_sent", "emails_sent", "emails_sent_with_copilot",
    "meeting_hours_summarized", "meetings_summarized",
    "chat_web_prompts", "chat_work_prompts",
]

_COL_WIDTHS = [36, 12, 20, 10, 14, 12, 13, 15, 20, 20, 25, 22, 12, 13, 23, 24, 22, 18, 18]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)
    for col, width in enumerate(_COL_WIDTHS, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width

    if not rows:
        ws.cell(row=2, column=1, value="— No Copilot Impact data imported (set VIVA_REPORT_IMPACT) —")
        return

    for r in rows:
        ws.append([r.get(f) for f in _FIELDS])
