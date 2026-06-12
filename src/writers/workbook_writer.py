from pathlib import Path

import openpyxl

from src.writers import (
    sheet_agents,
    sheet_ai_usage,
    sheet_az_health,
    sheet_connectors,
    sheet_crossref,
    sheet_dlp,
    sheet_environments,
    sheet_invocations,
    sheet_kpi_history,
    sheet_m365_admin_inventory,
    sheet_m365_copilot,
    sheet_m365_copilot_trend,
    sheet_m365_packages,
    sheet_m365_o365_users,
    sheet_m365_app_users,
    sheet_m365_usage_agents,
    sheet_m365_usage_agent_users,
    sheet_publishers,
    sheet_summary,
    sheet_teams_usage,
    sheet_viva,
    sheet_viva_adoption,
    sheet_viva_impact,
    sheet_viva_sessions,
    sheet_viva_topics,
    sheet_viva_wau,
    sheet_viva_autonomous,
)


def build_workbook(
    events: list[dict],
    connector_calls: list[dict],
    output_path: str,
    agents: list[dict] | None = None,
    environments: list[dict] | None = None,
    publishers: list[dict] | None = None,
    dlp_policies: list[dict] | None = None,
    agent_solutions: list[dict] | None = None,
    aad_users: dict[str, dict] | None = None,
    model_calls: list[dict] | None = None,
    health_detail: list[dict] | None = None,
    crossref_summary: list[dict] | None = None,
    copilot_usage: list[dict] | None = None,
    teams_usage: list[dict] | None = None,
    kpi_snapshots: list[dict] | None = None,
    viva_person_insights: list[dict] | None = None,
    viva_reports_cs_session_metrics: list[dict] | None = None,
    viva_reports_cs_topic_metrics: list[dict] | None = None,
    viva_reports_cs_weekly_active_users: list[dict] | None = None,
    viva_reports_cs_autonomous_metrics: list[dict] | None = None,
    viva_reports_cs_copilot_agents: dict[str, dict] | None = None,
    copilot_count_summary: list[dict] | None = None,
    copilot_count_trend: list[dict] | None = None,
    copilot_packages: list[dict] | None = None,
    o365_active_users: list[dict] | None = None,
    m365_app_users: list[dict] | None = None,
    viva_reports_copilot_adoption: list[dict] | None = None,
    viva_reports_copilot_impact: list[dict] | None = None,
    m365_admin_agent_inventory: list[dict] | None = None,
    m365_usage_agents: list[dict] | None = None,
    m365_usage_agent_users: list[dict] | None = None,
) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    latest_kpi = kpi_snapshots[0] if kpi_snapshots else None
    sheet_kpi_history.write(wb.create_sheet("KPI History"), kpi_snapshots or [])
    sheet_summary.write(
        wb.create_sheet("Summary"), events, connector_calls, model_calls or [],
        kpi_snapshot=latest_kpi,
        viva_reports_cs_sessions=viva_reports_cs_session_metrics or [],
        viva_reports_cs_wau=viva_reports_cs_weekly_active_users or [],
        viva_reports_cs_autonomous=viva_reports_cs_autonomous_metrics or [],
        viva_reports_cs_agents=viva_reports_cs_copilot_agents or {},
    )
    sheet_invocations.write(wb.create_sheet("Invocations"), events, connector_calls)
    sheet_connectors.write(wb.create_sheet("Connectors"), connector_calls)
    sheet_ai_usage.write(wb.create_sheet("AI_Model_Calls"), model_calls or [])
    sheet_agents.write(wb.create_sheet("Agents"), agents or [], environments or [], agent_solutions or [], aad_users or {})
    sheet_environments.write(wb.create_sheet("Environments"), environments or [])
    sheet_publishers.write(wb.create_sheet("Publishers"), publishers or [])
    sheet_dlp.write(wb.create_sheet("DLP Policies"), dlp_policies or [])
    sheet_m365_copilot.write(wb.create_sheet("M365_Copilot_Usage"), copilot_usage or [])
    sheet_m365_copilot_trend.write(wb.create_sheet("M365_Copilot_Trend"),
                                   copilot_count_summary or [], copilot_count_trend or [])
    sheet_m365_packages.write(wb.create_sheet("M365_Copilot_Packages"), copilot_packages or [])
    sheet_m365_o365_users.write(wb.create_sheet("M365_O365_Users"), o365_active_users or [])
    sheet_m365_app_users.write(wb.create_sheet("M365_App_Users"), m365_app_users or [])
    sheet_teams_usage.write(wb.create_sheet("Teams_Usage"), teams_usage or [])
    sheet_viva.write(wb.create_sheet("Viva_Person_Insights"), viva_person_insights or [], aad_users or {})
    sheet_viva_sessions.write(wb.create_sheet("Viva_CS_Sessions"), viva_reports_cs_session_metrics or [], viva_reports_cs_copilot_agents or {})
    sheet_viva_topics.write(wb.create_sheet("Viva_CS_Topics"), viva_reports_cs_topic_metrics or [], viva_reports_cs_copilot_agents or {})
    sheet_viva_wau.write(wb.create_sheet("Viva_CS_WAU"), viva_reports_cs_weekly_active_users or [], viva_reports_cs_copilot_agents or {})
    sheet_viva_autonomous.write(wb.create_sheet("Viva_CS_Autonomous"), viva_reports_cs_autonomous_metrics or [], viva_reports_cs_copilot_agents or {})
    sheet_viva_adoption.write(wb.create_sheet("Viva_Copilot_Adoption"), viva_reports_copilot_adoption or [])
    sheet_viva_impact.write(wb.create_sheet("Viva_Copilot_Impact"), viva_reports_copilot_impact or [])
    sheet_m365_admin_inventory.write(wb.create_sheet("M365_Agent_Inventory"), m365_admin_agent_inventory or [])
    sheet_m365_usage_agents.write(wb.create_sheet("M365_Usage_Agents"), m365_usage_agents or [])
    sheet_m365_usage_agent_users.write(wb.create_sheet("M365_Usage_AgentUsers"), m365_usage_agent_users or [])
    sheet_az_health.write(wb.create_sheet("AzureMonitor_Health"), health_detail or [])
    sheet_crossref.write(wb.create_sheet("CrossRef_Summary"), crossref_summary or [])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Saved: {output_path}")
