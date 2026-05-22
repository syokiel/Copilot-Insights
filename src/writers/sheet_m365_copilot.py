from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "User Principal Name",
    "Display Name",
    "Last Activity Date",
    "Teams Chats",
    "Teams Meetings",
    "Word",
    "Excel",
    "PowerPoint",
    "Outlook",
    "OneNote",
    "Loop",
    "Copilot Chat",
    "Report Period",
    "Report Refresh Date",
]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— Requires Reports.Read.All permission on the sync service principal —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r.get("user_principal_name", ""))
            ws.cell(row=i, column=2, value=r.get("display_name", ""))
            ws.cell(row=i, column=3, value=r.get("last_activity_date", ""))
            ws.cell(row=i, column=4, value=r.get("teams_chats"))
            ws.cell(row=i, column=5, value=r.get("teams_meetings"))
            ws.cell(row=i, column=6, value=r.get("word"))
            ws.cell(row=i, column=7, value=r.get("excel"))
            ws.cell(row=i, column=8, value=r.get("powerpoint"))
            ws.cell(row=i, column=9, value=r.get("outlook"))
            ws.cell(row=i, column=10, value=r.get("onenote"))
            ws.cell(row=i, column=11, value=r.get("loop"))
            ws.cell(row=i, column=12, value=r.get("copilot_chat"))
            ws.cell(row=i, column=13, value=r.get("report_period", ""))
            ws.cell(row=i, column=14, value=r.get("report_refresh_date", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
