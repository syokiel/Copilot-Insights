from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Conversation ID",
    "Agent ID",
    "Agent Name",
    "Env ID",
    "Invocation Time",
    "Dependency Name",
    "Result Code",
    "Dep Success",
    "Exception Type",
    "Exception Message",
    "Duration (ms)",
    "Alert Fired",
    "Alert Name",
]

_RED = PatternFill("solid", fgColor="FFCCCC")
_ORANGE = PatternFill("solid", fgColor="FFE5CC")


def write(ws: Worksheet, health_detail: list[dict]) -> None:
    write_headers(ws, HEADERS)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    if not health_detail:
        ws.cell(row=2, column=1, value="— No Azure Monitor data synced yet —")
        autofit_columns(ws, HEADERS)
        return

    for row_idx, row in enumerate(health_detail, start=2):
        dep_success = row.get("dependency_success")
        alert_fired = bool(row.get("alert_fired"))

        ws.cell(row=row_idx, column=1, value=row.get("conversation_id", ""))
        ws.cell(row=row_idx, column=2, value=row.get("agent_id", ""))
        ws.cell(row=row_idx, column=3, value=row.get("agent_name", ""))
        ws.cell(row=row_idx, column=4, value=row.get("env_id", ""))
        ws.cell(row=row_idx, column=5, value=row.get("invocation_time", ""))
        ws.cell(row=row_idx, column=6, value=row.get("dependency_name", ""))
        ws.cell(row=row_idx, column=7, value=row.get("dependency_result_code", ""))
        ws.cell(row=row_idx, column=8, value="No" if dep_success is False else ("Yes" if dep_success else ""))
        ws.cell(row=row_idx, column=9, value=row.get("exception_type", ""))
        ws.cell(row=row_idx, column=10, value=row.get("exception_message", ""))
        ws.cell(row=row_idx, column=11, value=row.get("duration_ms"))
        ws.cell(row=row_idx, column=12, value="Yes" if alert_fired else "No")
        ws.cell(row=row_idx, column=13, value=row.get("alert_name", ""))

        apply_row_style(ws, row_idx, len(HEADERS))

        # Conditional formatting: red for failed dependency, orange for alert
        if dep_success is False:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = _RED
        elif alert_fired:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col).fill = _ORANGE

    autofit_columns(ws, HEADERS)
