from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Title ID", "Name", "Status", "Channel", "Platform",
    "Owner", "Publisher", "Publisher Type", "Version",
    "Date Created", "Last Modified",
    "Can Read OD/SP", "Can Extend Graph", "Can Generate Images",
    "Can Use Code Interpreter", "Contains Uploaded Files",
    "Custom Actions", "Groups Shared", "Users Shared",
]

_BOOL_COLS = {
    "can_read_od_sp", "can_extend_graph", "can_generate_images",
    "can_use_code_interpreter", "contains_uploaded_files",
}


def _yn(v) -> str:
    return "Yes" if v else ("No" if v is not None else "")


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— No M365 Admin agent inventory imported (set M365ADMIN_AGENT_INVENTORY) —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1,  value=r.get("title_id", ""))
            ws.cell(row=i, column=2,  value=r.get("name", ""))
            ws.cell(row=i, column=3,  value=r.get("status", ""))
            ws.cell(row=i, column=4,  value=r.get("channel", ""))
            ws.cell(row=i, column=5,  value=r.get("platform", ""))
            ws.cell(row=i, column=6,  value=r.get("owner", ""))
            ws.cell(row=i, column=7,  value=r.get("publisher", ""))
            ws.cell(row=i, column=8,  value=r.get("publisher_type", ""))
            ws.cell(row=i, column=9,  value=r.get("version", ""))
            ws.cell(row=i, column=10, value=r.get("date_created", ""))
            ws.cell(row=i, column=11, value=r.get("last_modified", ""))
            ws.cell(row=i, column=12, value=_yn(r.get("can_read_od_sp")))
            ws.cell(row=i, column=13, value=_yn(r.get("can_extend_graph")))
            ws.cell(row=i, column=14, value=_yn(r.get("can_generate_images")))
            ws.cell(row=i, column=15, value=_yn(r.get("can_use_code_interpreter")))
            ws.cell(row=i, column=16, value=_yn(r.get("contains_uploaded_files")))
            ws.cell(row=i, column=17, value=r.get("custom_actions"))
            ws.cell(row=i, column=18, value=r.get("groups_shared", ""))
            ws.cell(row=i, column=19, value=r.get("users_shared", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
