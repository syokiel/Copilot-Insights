from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Agent ID", "Agent Name", "Username",
    "Creator Type", "Responses Sent", "Last Activity Date",
]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— No M365 Usage per-user data imported (set M365USAGE_REPORT_AgentUser) —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r.get("agent_id", ""))
            ws.cell(row=i, column=2, value=r.get("agent_name", ""))
            ws.cell(row=i, column=3, value=r.get("username", ""))
            ws.cell(row=i, column=4, value=r.get("creator_type", ""))
            ws.cell(row=i, column=5, value=r.get("responses_sent"))
            ws.cell(row=i, column=6, value=r.get("last_activity_date", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
