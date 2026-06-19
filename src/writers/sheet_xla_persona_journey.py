from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "Persona", "Journey", "Agents",
    "Total Sessions", "Resolved", "Escalated", "Abandoned", "Engaged",
    "Completion %", "Escalation %", "Abandonment %", "XLA Score",
]

_FIELDS = [
    "persona_type", "journey_name", "agent_count",
    "total_sessions", "resolved_sessions", "escalated_sessions",
    "abandoned_sessions", "engaged_sessions",
    "completion_rate_pct", "escalation_rate_pct", "abandonment_rate_pct", "xla_score",
]

_PERSONA_LABELS = {
    "end_user":        "End User",
    "it_support":      "IT Support",
    "hr":              "HR",
    "knowledge_worker": "Knowledge Worker",
    "operations":      "Operations",
}

_XLA_GOOD  = "FF00B050"  # green  ≥ 75
_XLA_WARN  = "FFFFCC00"  # amber  50–74
_XLA_BAD   = "FFFF0000"  # red    < 50


def _xla_fill(score) -> PatternFill | None:
    if score is None:
        return None
    s = float(score)
    colour = _XLA_GOOD if s >= 75 else (_XLA_WARN if s >= 50 else _XLA_BAD)
    return PatternFill("solid", fgColor=colour)


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— No XLA data: set AGENT_JOURNEY_MAP and ensure Viva CS session metrics are imported —")
        return

    xla_col = HEADERS.index("XLA Score") + 1

    for i, r in enumerate(rows, start=2):
        for col, field in enumerate(_FIELDS, start=1):
            val = r.get(field)
            if field == "persona_type":
                val = _PERSONA_LABELS.get(val, val)
            ws.cell(row=i, column=col, value=val)
        apply_row_style(ws, i, len(HEADERS))

        fill = _xla_fill(r.get("xla_score"))
        if fill:
            cell = ws.cell(row=i, column=xla_col)
            cell.fill = fill
            cell.font = Font(bold=True)

    autofit_columns(ws, HEADERS)
