from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import HEADER_FILL, HEADER_FONT, CENTER, write_headers

HEADERS = [
    "Person ID", "Metric Date", "Organization",
    "Total Actions", "Active Days", "Enabled Days",
    "Chat Work Prompts", "Chat Web Prompts",
    "Word (Work)", "Word (Web)",
    "Excel (Work)", "Excel (Web)",
    "PPT (Work)", "PPT (Web)",
    "Actions: Copilot Chat", "Actions: Excel",
    "Actions: Outlook", "Actions: PowerPoint",
    "Actions: Teams", "Actions: Word",
    "Meeting Hours Summarized",
]

_FIELDS = [
    "person_id", "metric_date", "organization",
    "total_copilot_actions", "total_copilot_active_days", "total_copilot_enabled_days",
    "chat_work_prompts", "chat_web_prompts",
    "word_work_prompts", "word_web_prompts",
    "excel_work_prompts", "excel_web_prompts",
    "ppt_work_prompts", "ppt_web_prompts",
    "actions_copilot_chat", "actions_excel",
    "actions_outlook", "actions_powerpoint",
    "actions_teams", "actions_word",
    "meeting_hours_summarized",
]

_COL_WIDTHS = [36, 12, 20, 14, 12, 13, 18, 16, 12, 12, 12, 12, 12, 12, 22, 16, 17, 22, 16, 16, 24]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)
    for col, width in enumerate(_COL_WIDTHS, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width

    if not rows:
        ws.cell(row=2, column=1, value="— No Copilot Adoption data imported (set VIVA_REPORT_ADOPTION) —")
        return

    # Use append for performance with large row counts (300K+)
    for r in rows:
        ws.append([r.get(f) for f in _FIELDS])
