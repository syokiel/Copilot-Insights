from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import LEFT, autofit_columns, write_headers

_SECTION_FILL = PatternFill("solid", fgColor="C6EFCE")
_SECTION_FONT = Font(bold=True, size=10, color="375623")

HEADERS = [
    "Snapshot Date",
    "Lookback Days",
    # M365 Copilot
    "Total Licenses",
    "Enabled Users",
    "Active Users",
    "Activation Rate",
    "Adoption Rate",
    "Power Users",
    "Total Prompts",
    "Avg Prompts / User",
    # Workload volumes
    "Copilot Chat Prompts",
    "Teams Prompts",
    "Outlook Prompts",
    "Excel Prompts",
    "Word Prompts",
    "PowerPoint Prompts",
    "OneNote Prompts",
    "Loop Prompts",
    # Studio agent adoption
    "Agent Adopters",
    "Agent Adoption %",
    # Agent inventory
    "Total Agents",
    "Active Agents",
    "Utilization Rate",
    "Production Agents",
    "Non-Prod Agents",
    "Agents with Owner",
    "Ownership %",
    "Total Conversations",
    # Environment breakdown
    "Env: Default",
    "Env: Developer",
    "Env: Teams",
    "Env: Production",
    "Env: Sandbox",
    "Env: Trial",
]

# Section group labels: (label, first_col, last_col) — 1-based
_SECTIONS = [
    ("",                   1,  2),
    ("M365 Copilot",       3,  10),
    ("Workloads",         11,  18),
    ("Agent Adoption",    19,  20),
    ("Agent Inventory",   21,  28),
    ("Environments",      29,  34),
]

# Columns that hold rates stored as a fraction (divide the % value by 100)
_PCT_COLS = {6, 7, 20, 23, 27}


def write(ws: Worksheet, snapshots: list[dict]) -> None:
    # Row 1 — section group labels
    for label, col_start, col_end in _SECTIONS:
        if label:
            cell = ws.cell(row=1, column=col_start, value=label)
            cell.fill = _SECTION_FILL
            cell.font = _SECTION_FONT
            cell.alignment = LEFT
            if col_end > col_start:
                ws.merge_cells(
                    start_row=1, start_column=col_start,
                    end_row=1, end_column=col_end,
                )

    # Row 2 — column headers
    write_headers(ws, HEADERS, start_row=2)
    ws.freeze_panes = "A3"  # override: keep both row 1 (section labels) and row 2 (headers) pinned

    if not snapshots:
        ws.cell(row=3, column=1, value="— No KPI snapshots yet — run sync to create the first one —")
        return

    def _pct(v):
        return (v / 100) if v is not None else None

    for row_idx, snap in enumerate(snapshots, start=3):
        vals = [
            (snap.get("snapshot_date") or "")[:19],
            snap.get("lookback_days"),
            snap.get("total_licenses"),
            snap.get("enabled_users"),
            snap.get("active_users"),
            _pct(snap.get("activation_rate")),
            _pct(snap.get("adoption_rate")),
            snap.get("power_users"),
            snap.get("total_prompts"),
            snap.get("avg_prompts_per_user"),
            snap.get("prompts_copilot_chat"),
            snap.get("prompts_teams"),
            snap.get("prompts_outlook"),
            snap.get("prompts_excel"),
            snap.get("prompts_word"),
            snap.get("prompts_powerpoint"),
            snap.get("prompts_onenote"),
            snap.get("prompts_loop"),
            snap.get("agent_adopters"),
            _pct(snap.get("agent_adoption_pct")),
            snap.get("total_agents"),
            snap.get("active_agents"),
            _pct(snap.get("utilization_rate")),
            snap.get("production_agents"),
            snap.get("non_prod_agents"),
            snap.get("agents_with_owner"),
            _pct(snap.get("ownership_pct")),
            snap.get("total_conversations"),
            snap.get("env_default"),
            snap.get("env_developer"),
            snap.get("env_teams"),
            snap.get("env_production"),
            snap.get("env_sandbox"),
            snap.get("env_trial"),
        ]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in _PCT_COLS:
                cell.number_format = "0.0%"

    autofit_columns(ws, HEADERS)
