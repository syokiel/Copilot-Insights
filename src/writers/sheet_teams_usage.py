from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import apply_row_style, autofit_columns, write_headers

HEADERS = [
    "User Principal Name",
    "Last Activity Date",
    "Team Chat Messages",
    "Private Chat Messages",
    "Calls",
    "Meetings",
    "Meetings Organized",
    "Meetings Attended",
    "Report Period",
    "Report Refresh Date",
]


def write(ws: Worksheet, rows: list[dict]) -> None:
    write_headers(ws, HEADERS)

    if not rows:
        ws.cell(row=2, column=1, value="— Requires Reports.Read.All permission on the sync service principal —")
    else:
        for i, r in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=r.get("user_principal_name", ""))
            ws.cell(row=i, column=2, value=r.get("last_activity_date", ""))
            ws.cell(row=i, column=3, value=r.get("team_chat_messages"))
            ws.cell(row=i, column=4, value=r.get("private_chat_messages"))
            ws.cell(row=i, column=5, value=r.get("calls"))
            ws.cell(row=i, column=6, value=r.get("meetings"))
            ws.cell(row=i, column=7, value=r.get("meetings_organized"))
            ws.cell(row=i, column=8, value=r.get("meetings_attended"))
            ws.cell(row=i, column=9, value=r.get("report_period", ""))
            ws.cell(row=i, column=10, value=r.get("report_refresh_date", ""))
            apply_row_style(ws, i, len(HEADERS))

    autofit_columns(ws, HEADERS)
