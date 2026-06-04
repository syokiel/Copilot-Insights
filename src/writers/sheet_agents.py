from openpyxl.worksheet.worksheet import Worksheet

from src.writers._style import autofit_columns, write_headers

_OWNER_NOT_FOUND = (
    "Owner information for these agents could not be found in the user directory. "
    "This may indicate the agents are not linked to a user account, "
    "or the accounts are not present in the directory."
)

HEADERS = [
    "Agent ID",
    "Agent Name",
    "AI Model",
    "Environment ID",
    "Environment Name",
    "Created Date",
    "Modified Date",
    "Published",
    "Created By",
    "Owner",
    "Created In",
    "Solution Name",
    "Solution Version",
    "Managed",
]


def write(
    ws: Worksheet,
    agents: list[dict],
    environments: list[dict] = [],
    agent_solutions: list[dict] = [],
    aad_users: dict[str, dict] = {},
) -> None:
    write_headers(ws, HEADERS)

    env_by_id = {e["environment_id"]: e["display_name"] for e in environments if e.get("environment_id")}
    sol_by_agent = {s["agent_id"]: s for s in agent_solutions if s.get("agent_id")}

    if not agents:
        ws.cell(row=2, column=1, value="— No agent data synced yet —")
    else:
        for row_idx, agent in enumerate(agents, start=2):
            agent_id = agent.get("agent_id") or agent.get("id", "")
            env_id = agent.get("environment_id") or agent.get("environmentId", "")
            sol = sol_by_agent.get(agent_id, {})
            published = agent.get("published_at") or agent.get("publishedDateTime", "")

            owner_id = agent.get("owner_id", "")
            if not owner_id:
                owner_display = ""
            elif owner_id in aad_users:
                user = aad_users[owner_id]
                owner_display = (
                    user.get("display_name") or user.get("upn", owner_id)
                    if user.get("found")
                    else _OWNER_NOT_FOUND
                )
            else:
                owner_display = owner_id  # not yet resolved

            ws.cell(row=row_idx, column=1, value=agent_id)
            ws.cell(row=row_idx, column=2, value=agent.get("display_name") or agent.get("name", ""))
            ws.cell(row=row_idx, column=3, value=agent.get("ai_model", ""))
            ws.cell(row=row_idx, column=4, value=env_id)
            ws.cell(row=row_idx, column=5, value=env_by_id.get(env_id, ""))
            ws.cell(row=row_idx, column=6, value=agent.get("created_at") or agent.get("createdDateTime", ""))
            ws.cell(row=row_idx, column=7, value=agent.get("modified_at") or agent.get("modifiedDateTime", ""))
            ws.cell(row=row_idx, column=8, value="Yes" if published else "No")
            ws.cell(row=row_idx, column=9, value=agent.get("created_by", ""))
            ws.cell(row=row_idx, column=10, value=owner_display)
            ws.cell(row=row_idx, column=11, value=agent.get("created_in", ""))
            ws.cell(row=row_idx, column=12, value=sol.get("solution_name", ""))
            ws.cell(row=row_idx, column=13, value=sol.get("version", ""))
            ws.cell(row=row_idx, column=14, value="Yes" if sol.get("is_managed") else ("No" if sol else ""))

    autofit_columns(ws, HEADERS)
